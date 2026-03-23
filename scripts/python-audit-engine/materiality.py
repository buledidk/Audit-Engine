"""
Materiality Calculator — ISA 320 / ISA 450
Reads a trial balance CSV or Excel file and calculates:
  - Overall materiality (ISA 320.10)
  - Performance materiality (ISA 320.11)
  - Trivial threshold (ISA 450.A2)

Usage:
  python scripts/materiality.py --client eim-learning-uk-ltd
  python scripts/materiality.py --client huge-uk-ltd --benchmark revenue
"""

import argparse
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from datetime import date

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


# ── Configuration ──────────────────────────────────────────────
BENCHMARKS = {
    "revenue": {"percentage": Decimal("0.02"), "description": "2% of Revenue (ISA 320.A4)"},
    "total_assets": {"percentage": Decimal("0.01"), "description": "1% of Total Assets (ISA 320.A4)"},
    "profit": {"percentage": Decimal("0.05"), "description": "5% of Profit Before Tax (ISA 320.A4)"},
    "expenses": {"percentage": Decimal("0.02"), "description": "2% of Total Expenses"},
}

PERFORMANCE_MATERIALITY_RATE = Decimal("0.75")  # 75% of overall (ISA 320.11)
TRIVIAL_THRESHOLD_RATE = Decimal("0.05")         # 5% of overall (ISA 450.A2)


def find_trial_balance(client_dir: Path) -> Path:
    """Find the trial balance file in the client's trial-balance folder."""
    tb_dir = client_dir / "trial-balance"
    if not tb_dir.exists():
        raise FileNotFoundError(f"No trial-balance folder found at {tb_dir}")

    for ext in ["*.xlsx", "*.xls", "*.csv"]:
        files = list(tb_dir.glob(ext))
        if files:
            return files[0]

    raise FileNotFoundError(f"No trial balance file found in {tb_dir}")


def load_trial_balance(filepath: Path) -> pd.DataFrame:
    """Load trial balance from CSV or Excel."""
    if filepath.suffix == ".csv":
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    # Standardise column names
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    return df


def extract_benchmark_value(df: pd.DataFrame, benchmark: str) -> Decimal:
    """
    Extract the benchmark value from the trial balance.
    Attempts to identify the relevant total from common column patterns.
    """
    # Common column name patterns
    amount_cols = [c for c in df.columns if any(
        term in c for term in ["amount", "balance", "total", "value", "debit", "credit"]
    )]

    if not amount_cols:
        raise ValueError(
            f"Cannot identify amount columns in trial balance. "
            f"Columns found: {list(df.columns)}"
        )

    account_col = None
    for candidate in ["account", "account_name", "description", "name", "nominal"]:
        if candidate in df.columns:
            account_col = candidate
            break

    if account_col is None:
        print("WARNING: No account name column found. Using total of all amounts.")
        total = df[amount_cols[0]].sum()
        return Decimal(str(total)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    # Try to find relevant accounts based on benchmark type
    search_terms = {
        "revenue": ["revenue", "turnover", "sales", "income"],
        "total_assets": ["total assets", "asset"],
        "profit": ["profit", "surplus", "net income"],
        "expenses": ["expense", "cost", "overhead"],
    }

    terms = search_terms.get(benchmark, ["total"])
    mask = df[account_col].astype(str).str.lower().str.contains("|".join(terms), na=False)

    if mask.any():
        total = df.loc[mask, amount_cols[0]].sum()
    else:
        print(f"WARNING: Could not find {benchmark} accounts. Using total of column '{amount_cols[0]}'.")
        total = df[amount_cols[0]].abs().sum()

    return Decimal(str(abs(total))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_materiality(
    benchmark_value: Decimal,
    benchmark_name: str,
) -> dict:
    """Calculate all materiality levels per ISA 320."""
    config = BENCHMARKS[benchmark_name]

    overall = (benchmark_value * config["percentage"]).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    performance = (overall * PERFORMANCE_MATERIALITY_RATE).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    trivial = (overall * TRIVIAL_THRESHOLD_RATE).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    return {
        "benchmark_name": benchmark_name,
        "benchmark_description": config["description"],
        "benchmark_value": benchmark_value,
        "percentage": config["percentage"],
        "overall_materiality": overall,
        "performance_materiality": performance,
        "performance_materiality_rate": PERFORMANCE_MATERIALITY_RATE,
        "trivial_threshold": trivial,
        "trivial_threshold_rate": TRIVIAL_THRESHOLD_RATE,
    }


def generate_workpaper(results: dict, client_name: str, output_path: Path) -> None:
    """Generate formatted Excel workpaper for materiality calculation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiality"

    # ── Styles ──
    title_font = Font(name="Arial", size=14, bold=True, color="0D1B2A")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    label_font = Font(name="Arial", size=11, bold=True, color="0D1B2A")
    value_font = Font(name="Arial", size=11)
    isa_font = Font(name="Arial", size=10, italic=True, color="6B7280")
    result_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # ── Column widths ──
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    # ── Header section ──
    row = 1
    ws.cell(row=row, column=1, value="MATERIALITY CALCULATION WORKPAPER").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Client: {client_name.replace('-', ' ').title()}").font = label_font
    row += 1
    ws.cell(row=row, column=1, value=f"Prepared by: DK (Senior Auditor)").font = value_font
    ws.cell(row=row, column=3, value=f"Date: {date.today().strftime('%d %B %Y')}").font = value_font
    row += 1
    ws.cell(row=row, column=1, value="Standard: ISA 320 (Materiality in Planning and Performing an Audit)").font = isa_font
    row += 2

    # ── Objective ──
    ws.cell(row=row, column=1, value="OBJECTIVE").font = label_font
    row += 1
    ws.cell(row=row, column=1, value="To determine materiality levels for the audit in accordance with ISA 320.").font = value_font
    row += 2

    # ── Calculation table ──
    headers = ["Item", "Value", "Rate", "ISA Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    row += 1

    # Data rows
    data = [
        ("Benchmark Selected", results["benchmark_description"], "", "ISA 320.A4"),
        ("Benchmark Value", f"GBP {results['benchmark_value']:,.2f}", "", ""),
        ("Overall Materiality", f"GBP {results['overall_materiality']:,.2f}",
         f"{results['percentage'] * 100:.1f}%", "ISA 320.10"),
        ("Performance Materiality", f"GBP {results['performance_materiality']:,.2f}",
         f"{results['performance_materiality_rate'] * 100:.0f}% of overall", "ISA 320.11"),
        ("Trivial Threshold", f"GBP {results['trivial_threshold']:,.2f}",
         f"{results['trivial_threshold_rate'] * 100:.0f}% of overall", "ISA 450.A2"),
    ]

    for item, value, rate, ref in data:
        ws.cell(row=row, column=1, value=item).font = label_font
        ws.cell(row=row, column=2, value=value).font = value_font
        ws.cell(row=row, column=3, value=rate).font = value_font
        ws.cell(row=row, column=4, value=ref).font = isa_font
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            if item in ("Overall Materiality", "Performance Materiality", "Trivial Threshold"):
                ws.cell(row=row, column=col).fill = result_fill
        row += 1

    row += 1

    # ── Conclusion ──
    ws.cell(row=row, column=1, value="CONCLUSION").font = label_font
    row += 1
    ws.cell(
        row=row, column=1,
        value=(
            f"Based on the benchmark of {results['benchmark_description'].lower()}, "
            f"overall materiality has been set at GBP {results['overall_materiality']:,.2f}. "
            f"Performance materiality is GBP {results['performance_materiality']:,.2f} "
            f"and the trivial threshold is GBP {results['trivial_threshold']:,.2f}. "
            f"These levels are considered appropriate for the nature and circumstances "
            f"of the entity."
        ),
    ).font = value_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Workpaper saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Calculate audit materiality per ISA 320")
    parser.add_argument("--client", required=True, help="Client folder name (e.g. eim-learning-uk-ltd)")
    parser.add_argument("--benchmark", default="revenue", choices=list(BENCHMARKS.keys()),
                        help="Materiality benchmark (default: revenue)")
    parser.add_argument("--base-dir", default=".", help="Base project directory")
    args = parser.parse_args()

    base = Path(args.base_dir)
    client_dir = base / "engagements" / args.client

    if not client_dir.exists():
        print(f"ERROR: Client directory not found: {client_dir}")
        sys.exit(1)

    print(f"Client: {args.client}")
    print(f"Benchmark: {args.benchmark}")

    # Find and load trial balance
    tb_path = find_trial_balance(client_dir)
    print(f"Trial balance: {tb_path.name}")
    df = load_trial_balance(tb_path)
    print(f"Records loaded: {len(df)}")

    # Extract benchmark and calculate
    benchmark_value = extract_benchmark_value(df, args.benchmark)
    print(f"Benchmark value: GBP {benchmark_value:,.2f}")

    results = calculate_materiality(benchmark_value, args.benchmark)

    print(f"\n{'='*50}")
    print(f"Overall Materiality:      GBP {results['overall_materiality']:,.2f}")
    print(f"Performance Materiality:  GBP {results['performance_materiality']:,.2f}")
    print(f"Trivial Threshold:        GBP {results['trivial_threshold']:,.2f}")
    print(f"{'='*50}\n")

    # Generate workpaper
    output = client_dir / "working-papers" / "materiality.xlsx"
    generate_workpaper(results, args.client, output)


if __name__ == "__main__":
    main()
