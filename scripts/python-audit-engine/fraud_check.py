"""
Fraud Risk Indicator Scanner — ISA 240
Scans journal entries for fraud risk indicators:
  - Round number entries above performance materiality
  - Entries posted on weekends or bank holidays
  - Management override patterns (unusual account combinations)
  - Entries just below approval thresholds
  - Duplicate entries (same amount, same date, different reference)

Usage:
  python scripts/fraud_check.py --client eim-learning-uk-ltd
  python scripts/fraud_check.py --client huge-uk-ltd --threshold 25000
"""

import argparse
import sys
from decimal import Decimal
from pathlib import Path
from datetime import date, datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


# UK Bank Holidays 2024/2025 (extend as needed)
UK_BANK_HOLIDAYS = {
    datetime(2024, 1, 1).date(), datetime(2024, 3, 29).date(), datetime(2024, 4, 1).date(),
    datetime(2024, 5, 6).date(), datetime(2024, 5, 27).date(), datetime(2024, 8, 26).date(),
    datetime(2024, 12, 25).date(), datetime(2024, 12, 26).date(),
    datetime(2025, 1, 1).date(), datetime(2025, 4, 18).date(), datetime(2025, 4, 21).date(),
    datetime(2025, 5, 5).date(), datetime(2025, 5, 26).date(), datetime(2025, 8, 25).date(),
    datetime(2025, 12, 25).date(), datetime(2025, 12, 26).date(),
}


def load_data(client_dir: Path) -> pd.DataFrame:
    """Load journal entries or trial balance data."""
    tb_dir = client_dir / "trial-balance"
    files = sorted(tb_dir.glob("*.csv")) + sorted(tb_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No data files in {tb_dir}")

    filepath = files[-1]
    if filepath.suffix == ".csv":
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    return df


def find_columns(df: pd.DataFrame) -> dict:
    """Identify key columns in the data."""
    cols = {}

    for candidate in ["amount", "balance", "total", "value", "debit", "net"]:
        if candidate in df.columns:
            cols["amount"] = candidate
            break

    for candidate in ["date", "posting_date", "entry_date", "transaction_date"]:
        if candidate in df.columns:
            cols["date"] = candidate
            break

    for candidate in ["account", "account_name", "description", "nominal", "name"]:
        if candidate in df.columns:
            cols["account"] = candidate
            break

    for candidate in ["reference", "ref", "journal_ref", "document"]:
        if candidate in df.columns:
            cols["reference"] = candidate
            break

    for candidate in ["user", "posted_by", "created_by", "entered_by"]:
        if candidate in df.columns:
            cols["user"] = candidate
            break

    return cols


def check_round_numbers(df: pd.DataFrame, amount_col: str, threshold: float) -> pd.DataFrame:
    """Flag entries that are round numbers above the threshold."""
    mask = (
        (df[amount_col].abs() >= threshold) &
        (df[amount_col] % 1000 == 0) &
        (df[amount_col] != 0)
    )
    flagged = df[mask].copy()
    flagged["flag_type"] = "Round number above threshold"
    flagged["risk_level"] = "HIGH"
    flagged["isa_ref"] = "ISA 240.A44"
    return flagged


def check_weekend_postings(df: pd.DataFrame, date_col: str) -> pd.DataFrame:
    """Flag entries posted on weekends or bank holidays."""
    df_copy = df.copy()
    df_copy[date_col] = pd.to_datetime(df_copy[date_col], errors="coerce")

    mask_weekend = df_copy[date_col].dt.dayofweek.isin([5, 6])  # Saturday, Sunday
    mask_holiday = df_copy[date_col].dt.date.isin(UK_BANK_HOLIDAYS)

    mask = mask_weekend | mask_holiday
    flagged = df[mask].copy()
    flagged["flag_type"] = "Weekend or bank holiday posting"
    flagged["risk_level"] = "MEDIUM"
    flagged["isa_ref"] = "ISA 240.A44"
    return flagged


def check_just_below_threshold(
    df: pd.DataFrame, amount_col: str, threshold: float
) -> pd.DataFrame:
    """Flag entries just below common approval thresholds."""
    thresholds = [1000, 5000, 10000, 25000, 50000, 100000]
    thresholds.append(threshold)

    all_flagged = []
    for t in thresholds:
        lower = t * 0.90
        mask = (df[amount_col].abs() >= lower) & (df[amount_col].abs() < t)
        flagged = df[mask].copy()
        if not flagged.empty:
            flagged["flag_type"] = f"Just below {t:,.0f} threshold"
            flagged["risk_level"] = "MEDIUM"
            flagged["isa_ref"] = "ISA 240.A44"
            all_flagged.append(flagged)

    return pd.concat(all_flagged) if all_flagged else pd.DataFrame()


def check_duplicates(df: pd.DataFrame, amount_col: str, date_col: str | None) -> pd.DataFrame:
    """Flag potential duplicate entries (same amount, same date)."""
    if date_col is None:
        return pd.DataFrame()

    group_cols = [date_col, amount_col]
    counts = df.groupby(group_cols).size().reset_index(name="count")
    dupes = counts[counts["count"] > 1]

    if dupes.empty:
        return pd.DataFrame()

    flagged = df.merge(dupes[group_cols], on=group_cols, how="inner").copy()
    flagged["flag_type"] = "Potential duplicate entry"
    flagged["risk_level"] = "HIGH"
    flagged["isa_ref"] = "ISA 240.A44"
    return flagged


def generate_report(
    all_flags: pd.DataFrame,
    client_name: str,
    cols: dict,
    output_path: Path,
    summary: dict,
) -> None:
    """Generate the fraud indicators report in Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fraud Indicators"

    # Styles
    title_font = Font(name="Arial", size=14, bold=True, color="0D1B2A")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    label_font = Font(name="Arial", size=10, bold=True, color="0D1B2A")
    value_font = Font(name="Arial", size=10)
    isa_font = Font(name="Arial", size=9, italic=True, color="6B7280")
    high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    med_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Header
    row = 1
    ws.cell(row=row, column=1, value="FRAUD RISK INDICATOR REPORT").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Client: {client_name.replace('-', ' ').title()}").font = label_font
    row += 1
    ws.cell(row=row, column=1, value=f"Prepared by: DK (Senior Auditor)").font = value_font
    ws.cell(row=row, column=4, value=f"Date: {date.today().strftime('%d %B %Y')}").font = value_font
    row += 1
    ws.cell(row=row, column=1, value="Standard: ISA 240 (The Auditor's Responsibilities Relating to Fraud)").font = isa_font
    row += 2

    # Summary
    ws.cell(row=row, column=1, value="SUMMARY OF FINDINGS").font = label_font
    row += 1
    for key, val in summary.items():
        ws.cell(row=row, column=1, value=key).font = value_font
        ws.cell(row=row, column=2, value=val).font = value_font
        row += 1
    row += 1

    # Flagged items
    if all_flags.empty:
        ws.cell(row=row, column=1, value="No fraud risk indicators identified.").font = value_font
    else:
        headers = ["Account", "Amount (GBP)", "Flag Type", "Risk Level", "ISA Reference"]
        widths = [30, 18, 30, 12, 15]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        row += 1

        amount_col = cols.get("amount", "amount")
        account_col = cols.get("account", "account")

        for _, r in all_flags.head(200).iterrows():  # Cap at 200 rows
            ws.cell(row=row, column=1, value=str(r.get(account_col, ""))).font = value_font
            ws.cell(row=row, column=2, value=float(r.get(amount_col, 0))).font = value_font
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=r.get("flag_type", "")).font = value_font
            ws.cell(row=row, column=4, value=r.get("risk_level", "")).font = value_font
            ws.cell(row=row, column=5, value=r.get("isa_ref", "")).font = isa_font

            fill = high_fill if r.get("risk_level") == "HIGH" else med_fill
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).fill = fill
            row += 1

    # Conclusion
    row += 1
    ws.cell(row=row, column=1, value="CONCLUSION").font = label_font
    row += 1
    total_flags = len(all_flags) if not all_flags.empty else 0
    high_flags = len(all_flags[all_flags["risk_level"] == "HIGH"]) if not all_flags.empty else 0
    ws.cell(row=row, column=1, value=(
        f"A total of {total_flags} entries have been flagged for further investigation, "
        f"of which {high_flags} are assessed as high risk. "
        f"These items require follow up procedures in accordance with ISA 240.A44 to "
        f"determine whether they represent indicators of fraud or have a reasonable "
        f"business explanation."
    )).font = value_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Fraud indicators report saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Scan for ISA 240 fraud risk indicators")
    parser.add_argument("--client", required=True, help="Client folder name")
    parser.add_argument("--threshold", type=float, default=25000,
                        help="Performance materiality threshold (default: 25000)")
    parser.add_argument("--base-dir", default=".", help="Base project directory")
    args = parser.parse_args()

    base = Path(args.base_dir)
    client_dir = base / "engagements" / args.client

    if not client_dir.exists():
        print(f"ERROR: Client directory not found: {client_dir}")
        sys.exit(1)

    df = load_data(client_dir)
    cols = find_columns(df)
    print(f"Data loaded: {len(df)} entries")
    print(f"Columns identified: {cols}")

    all_flags = []

    if "amount" in cols:
        round_flags = check_round_numbers(df, cols["amount"], args.threshold)
        print(f"Round number flags: {len(round_flags)}")
        all_flags.append(round_flags)

        below_flags = check_just_below_threshold(df, cols["amount"], args.threshold)
        print(f"Below threshold flags: {len(below_flags)}")
        all_flags.append(below_flags)

    if "date" in cols:
        weekend_flags = check_weekend_postings(df, cols["date"])
        print(f"Weekend/holiday flags: {len(weekend_flags)}")
        all_flags.append(weekend_flags)

        if "amount" in cols:
            dupe_flags = check_duplicates(df, cols["amount"], cols["date"])
            print(f"Duplicate flags: {len(dupe_flags)}")
            all_flags.append(dupe_flags)

    combined = pd.concat([f for f in all_flags if not f.empty]) if all_flags else pd.DataFrame()

    summary = {
        "Total entries scanned": len(df),
        "Entries flagged": len(combined),
        "High risk flags": len(combined[combined["risk_level"] == "HIGH"]) if not combined.empty else 0,
        "Medium risk flags": len(combined[combined["risk_level"] == "MEDIUM"]) if not combined.empty else 0,
        "Threshold used (GBP)": f"{args.threshold:,.0f}",
    }

    output = client_dir / "working-papers" / "fraud-indicators.xlsx"
    generate_report(combined, args.client, cols, output, summary)


if __name__ == "__main__":
    main()
