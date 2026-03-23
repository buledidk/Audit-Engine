"""
Trial Balance Analytics — ISA 520 (Analytical Procedures)
Reads trial balance data and performs:
  - Year on year variance analysis
  - Ratio analysis (gross margin, current ratio, etc.)
  - Exception flagging based on materiality thresholds
  - Trend identification

Usage:
  python scripts/tb_analytics.py --client eim-learning-uk-ltd
"""

import argparse
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from datetime import date

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)


def find_files(client_dir: Path) -> tuple[Path, Path | None]:
    """Find current year TB and optionally prior year TB."""
    tb_dir = client_dir / "trial-balance"
    files = sorted(tb_dir.glob("*.csv")) + sorted(tb_dir.glob("*.xlsx")) + sorted(tb_dir.glob("*.xls"))

    if not files:
        raise FileNotFoundError(f"No trial balance files in {tb_dir}")

    current = files[-1]  # Most recent file
    prior = files[-2] if len(files) >= 2 else None

    return current, prior


def load_tb(filepath: Path) -> pd.DataFrame:
    """Load and standardise a trial balance file."""
    if filepath.suffix == ".csv":
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    return df


def identify_columns(df: pd.DataFrame) -> tuple[str, str]:
    """Identify the account name and amount columns."""
    account_col = None
    for candidate in ["account", "account_name", "description", "name", "nominal", "account_description"]:
        if candidate in df.columns:
            account_col = candidate
            break

    amount_col = None
    for candidate in ["amount", "balance", "total", "net", "value"]:
        if candidate in df.columns:
            amount_col = candidate
            break

    if account_col is None:
        account_col = df.columns[0]
    if amount_col is None:
        # Try to find numeric column
        numeric_cols = df.select_dtypes(include=["number"]).columns
        if len(numeric_cols) > 0:
            amount_col = numeric_cols[0]
        else:
            raise ValueError("Cannot identify a numeric amount column in the trial balance.")

    return account_col, amount_col


def variance_analysis(
    current_df: pd.DataFrame,
    prior_df: pd.DataFrame | None,
    account_col: str,
    amount_col: str,
    materiality: Decimal,
) -> pd.DataFrame:
    """Compare current to prior year and flag material variances."""
    result = current_df[[account_col, amount_col]].copy()
    result.columns = ["account", "current_year"]

    if prior_df is not None:
        p_account_col, p_amount_col = identify_columns(prior_df)
        prior = prior_df[[p_account_col, p_amount_col]].copy()
        prior.columns = ["account", "prior_year"]
        result = result.merge(prior, on="account", how="outer")
        result.fillna(0, inplace=True)
    else:
        result["prior_year"] = 0

    result["variance"] = result["current_year"] - result["prior_year"]
    result["variance_pct"] = result.apply(
        lambda r: (r["variance"] / r["prior_year"] * 100) if r["prior_year"] != 0 else 0, axis=1
    )
    result["material"] = result["variance"].abs() >= float(materiality)

    return result.sort_values("variance", key=abs, ascending=False)


def ratio_analysis(df: pd.DataFrame, account_col: str, amount_col: str) -> list[dict]:
    """Calculate key financial ratios from trial balance data."""
    ratios = []

    def find_total(search_terms: list[str]) -> float:
        mask = df[account_col].astype(str).str.lower().str.contains("|".join(search_terms), na=False)
        return float(df.loc[mask, amount_col].sum()) if mask.any() else 0.0

    revenue = abs(find_total(["revenue", "turnover", "sales"]))
    cost_of_sales = abs(find_total(["cost of sales", "cost of goods", "direct cost"]))
    total_assets = abs(find_total(["total asset", "fixed asset", "current asset"]))
    current_assets = abs(find_total(["current asset", "trade debtor", "cash", "bank", "stock", "inventory"]))
    current_liabilities = abs(find_total(["current liabilit", "trade creditor", "accrual", "vat"]))
    total_expenses = abs(find_total(["expense", "overhead", "admin", "cost"]))

    if revenue > 0 and cost_of_sales > 0:
        gm = ((revenue - cost_of_sales) / revenue) * 100
        ratios.append({"ratio": "Gross Margin %", "value": f"{gm:.1f}%",
                        "isa_ref": "ISA 520.A5", "comment": "Compare to prior year and industry"})

    if current_assets > 0 and current_liabilities > 0:
        cr = current_assets / current_liabilities
        ratios.append({"ratio": "Current Ratio", "value": f"{cr:.2f}",
                        "isa_ref": "ISA 520.A5", "comment": "Below 1.0 indicates liquidity risk"})

    if revenue > 0:
        expense_ratio = (total_expenses / revenue) * 100
        ratios.append({"ratio": "Expense to Revenue %", "value": f"{expense_ratio:.1f}%",
                        "isa_ref": "ISA 520.A5", "comment": "Monitor for unexpected increases"})

    if total_assets > 0 and revenue > 0:
        asset_turnover = revenue / total_assets
        ratios.append({"ratio": "Asset Turnover", "value": f"{asset_turnover:.2f}",
                        "isa_ref": "ISA 520.A5", "comment": "Efficiency of asset utilisation"})

    return ratios


def generate_workpaper(
    variances: pd.DataFrame,
    ratios: list[dict],
    client_name: str,
    output_path: Path,
    has_prior: bool,
) -> None:
    """Generate formatted analytics workpaper in Excel."""
    wb = Workbook()

    # ── Styles ──
    title_font = Font(name="Arial", size=14, bold=True, color="0D1B2A")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    label_font = Font(name="Arial", size=10, bold=True, color="0D1B2A")
    value_font = Font(name="Arial", size=10)
    isa_font = Font(name="Arial", size=9, italic=True, color="6B7280")
    flag_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # ── Sheet 1: Variance Analysis ──
    ws = wb.active
    ws.title = "Variance Analysis"

    row = 1
    ws.cell(row=row, column=1, value="ANALYTICAL PROCEDURES WORKPAPER").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Client: {client_name.replace('-', ' ').title()}").font = label_font
    row += 1
    ws.cell(row=row, column=1, value=f"Prepared by: DK (Senior Auditor)").font = value_font
    ws.cell(row=row, column=4, value=f"Date: {date.today().strftime('%d %B %Y')}").font = value_font
    row += 1
    ws.cell(row=row, column=1, value="Standard: ISA 520 (Analytical Procedures)").font = isa_font
    row += 1
    ws.cell(row=row, column=1, value="Objective: Identify unusual trends, variances, and relationships requiring further investigation.").font = value_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Headers
    headers = ["Account", "Current Year (GBP)", "Prior Year (GBP)", "Variance (GBP)", "Variance %", "Material?"]
    widths = [35, 18, 18, 18, 12, 10]
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    row += 1

    # Data
    for _, r in variances.iterrows():
        ws.cell(row=row, column=1, value=str(r["account"])).font = value_font
        ws.cell(row=row, column=2, value=float(r["current_year"])).font = value_font
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=float(r["prior_year"])).font = value_font
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(r["variance"])).font = value_font
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(r["variance_pct"])).font = value_font
        ws.cell(row=row, column=5).number_format = '0.0%'
        ws.cell(row=row, column=6, value="YES" if r["material"] else "No").font = value_font

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            if r["material"]:
                ws.cell(row=row, column=col).fill = flag_fill

        row += 1

    # ── Sheet 2: Ratio Analysis ──
    ws2 = wb.create_sheet("Ratio Analysis")
    row = 1
    ws2.cell(row=row, column=1, value="KEY RATIO ANALYSIS").font = title_font
    row += 2

    headers = ["Ratio", "Value", "ISA Reference", "Auditor Comment"]
    widths = [25, 15, 15, 40]
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = width
    row += 1

    for ratio in ratios:
        ws2.cell(row=row, column=1, value=ratio["ratio"]).font = label_font
        ws2.cell(row=row, column=2, value=ratio["value"]).font = value_font
        ws2.cell(row=row, column=3, value=ratio["isa_ref"]).font = isa_font
        ws2.cell(row=row, column=4, value=ratio["comment"]).font = value_font
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = border
        row += 1

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Analytics workpaper saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run ISA 520 analytical procedures")
    parser.add_argument("--client", required=True, help="Client folder name")
    parser.add_argument("--materiality", type=float, default=50000,
                        help="Overall materiality threshold in GBP (default: 50000)")
    parser.add_argument("--base-dir", default=".", help="Base project directory")
    args = parser.parse_args()

    base = Path(args.base_dir)
    client_dir = base / "engagements" / args.client
    materiality = Decimal(str(args.materiality))

    if not client_dir.exists():
        print(f"ERROR: Client directory not found: {client_dir}")
        sys.exit(1)

    current_path, prior_path = find_files(client_dir)
    print(f"Current year: {current_path.name}")
    print(f"Prior year: {prior_path.name if prior_path else 'Not available'}")

    current_df = load_tb(current_path)
    prior_df = load_tb(prior_path) if prior_path else None
    account_col, amount_col = identify_columns(current_df)

    # Variance analysis
    variances = variance_analysis(current_df, prior_df, account_col, amount_col, materiality)
    material_count = variances["material"].sum()
    print(f"\nAccounts analysed: {len(variances)}")
    print(f"Material variances flagged: {material_count}")

    # Ratio analysis
    ratios = ratio_analysis(current_df, account_col, amount_col)
    print(f"Ratios calculated: {len(ratios)}")

    # Generate workpaper
    output = client_dir / "working-papers" / "analytics.xlsx"
    generate_workpaper(variances, ratios, args.client, output, prior_df is not None)


if __name__ == "__main__":
    main()
